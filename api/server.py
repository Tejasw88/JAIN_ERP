from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
import os
import logging
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import pymysql
from contextlib import contextmanager
import io
import json
from openpyxl import load_workbook
import math
import random
import string

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# TiDB Connection Configuration
DB_CONFIG = {
    'host': os.environ.get('DB_HOST'),
    'port': int(os.environ.get('DB_PORT', 4000)),
    'user': os.environ.get('DB_USERNAME'),
    'password': os.environ.get('DB_PASSWORD'),
    'database': os.environ.get('DB_DATABASE'),
    'ssl': {'ssl': {}},  # TiDB Cloud SSL
    'cursorclass': pymysql.cursors.DictCursor
}

JWT_SECRET = os.environ.get('JWT_SECRET', 'default_secret_key')
JWT_ALGORITHM = "HS256"
JWT_EXPIRY_DAYS = 7

# Create the main app
app = FastAPI(title="Jain-Edu-Hub API", version="2.0.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Initialize Rate Limiter
def get_ip(request: Request):
    if request.client and request.client.host:
        return request.client.host
    return "127.0.0.1"

limiter = Limiter(key_func=get_ip)
app.state.limiter = limiter

def rate_limit_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(
        status_code=429,
        content={"detail": "Too many login attempts. Please try again later."}
    )

app.add_exception_handler(RateLimitExceeded, rate_limit_handler)
app.add_middleware(SlowAPIMiddleware)

security = HTTPBearer()

# Middleware for Security Headers
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self' 'unsafe-inline' 'unsafe-eval'; style-src 'self' 'unsafe-inline'; img-src 'self' data: https:; font-src 'self' data:; connect-src 'self' http://localhost:8000 https://*.onrender.com;"
        return response

app.add_middleware(SecurityHeadersMiddleware)

# CORS Configuration - Allow all origins for Vercel deployment
# Note: With credentials=False, we can use wildcard. Auth is via Bearer token in header.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,  # Must be False when using wildcard origin
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database connection helper
@contextmanager
def get_db_connection():
    connection = pymysql.connect(**DB_CONFIG)
    try:
        yield connection
    finally:
        connection.close()

def resolve_department_identifiers(dept_id: str, cursor) -> List[str]:
    """Given a department name or code, return a list containing both [name, code]"""
    if not dept_id or dept_id == 'all':
        return []
    cursor.execute("SELECT name, code FROM departments WHERE name = %s OR code = %s", (dept_id, dept_id))
    res = cursor.fetchone()
    if res:
        # Filter out None values and return unique identifiers
        return list(set(filter(None, [res['name'], res['code'], dept_id])))
    return [dept_id]


# Pydantic Models
class LoginRequest(BaseModel):
    identifier: str
    password: str

class LoginResponse(BaseModel):
    token: str
    user: dict

class UserCreate(BaseModel):
    username: str
    email: str
    password: Optional[str] = "123456789"
    role: str
    name: str
    idno: Optional[str] = None
    department: Optional[str] = None
    year: Optional[str] = None
    section: Optional[str] = "A"
    parent_id: Optional[int] = None

class SubjectTeacherPair(BaseModel):
    course_id: int
    teacher_id: int
    slots_per_week: int = 4

class TimetableGenerateRequest(BaseModel):
    department: str
    year: str
    section: str
    pairings: List[SubjectTeacherPair]

class DepartmentCreate(BaseModel):
    name: str
    code: str

class CourseCreate(BaseModel):
    name: str
    code: str
    department: Optional[str] = None
    year: Optional[str] = None

class GradeCreate(BaseModel):
    student_id: int
    course_id: int
    course_name: str
    title: str
    marks: int
    max_marks: int = 100

class AttendanceRecord(BaseModel):
    student_id: int
    status: str

class AttendanceCreate(BaseModel):
    course_id: int
    course_name: str
    department: Optional[str] = None
    year: Optional[str] = None
    date: str
    records: List[AttendanceRecord]

class ClassworkCreate(BaseModel):
    course_id: int
    department: Optional[str] = None
    year: Optional[str] = None
    type: str = "Assignment"
    title: str
    description: Optional[str] = None
    due_date: Optional[str] = None
    max_marks: int = 100

class SubmissionCreate(BaseModel):
    classwork_id: int
    content: Optional[str] = None

class ParentLinkRequest(BaseModel):
    parent_id: int
    student_username: str
    student_idno_digits: str

# JWT Helper Functions
def create_token(user_data: dict) -> str:
    payload = {
        'user_id': user_data['id'],
        'username': user_data['username'],
        'email': user_data.get('email', ''),
        'role': user_data['role'],
        'exp': datetime.now(timezone.utc) + timedelta(days=JWT_EXPIRY_DAYS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_role(*roles):
    def role_checker(token: dict = Depends(verify_token)):
        if token['role'] not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return token
    return role_checker

def validate_domain(email: str, role: str) -> bool:
    if role in ['Student', 'Teacher']:
        if not email.endswith('@jainuniversity.ac.in'):
            return False
    return True

# Routes
@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

@api_router.post("/auth/login", response_model=LoginResponse)
@limiter.limit("5/minute")
async def login(request: Request, login_data: LoginRequest):
    try:
        # Use request object for rate limiting
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT * FROM users WHERE username = %s OR email = %s
                """, (login_data.identifier, login_data.identifier))
                user = cursor.fetchone()
                
                if not user:
                    raise HTTPException(status_code=401, detail="Invalid credentials")
                
                stored_password = user.get('password', '')
                if not bcrypt.checkpw(login_data.password.encode('utf-8'), stored_password.encode('utf-8')):
                    raise HTTPException(status_code=401, detail="Invalid credentials")
                
                user_email = user.get('email', '')
                if user['role'] in ['Student', 'Teacher']:
                    if user_email and not user_email.endswith('@jainuniversity.ac.in'):
                        raise HTTPException(status_code=403, detail="Students and Teachers must use @jainuniversity.ac.in domain")
                
                token = create_token(user)
                
                user_response = {
                    'id': user['id'],
                    'username': user['username'],
                    'email': user.get('email', ''),
                    'role': user['role'],
                    'full_name': user.get('name', ''),
                    'usn': user.get('idno'),
                    'department': user.get('department'),
                    'year': user.get('year'),
                    'linked_student_id': user.get('parent_id'),
                    'must_change_password': user.get('must_change_password', False)
                }
                
                return {"token": token, "user": user_response}
    except Exception as e:
        detailed_error = f"{type(e).__name__}: {str(e)}"
        logging.error(f"LOGIN EXCEPTION: {detailed_error}")
        raise HTTPException(status_code=500, detail=detailed_error)


@api_router.get("/auth/me")
async def get_current_user(token: dict = Depends(verify_token)):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM users WHERE id = %s", (token['user_id'],))
            user = cursor.fetchone()
            if not user:
                raise HTTPException(status_code=404, detail="User not found")
            
            return {
                'id': user['id'],
                'username': user['username'],
                'email': user.get('email', ''),
                'role': user['role'],
                'full_name': user.get('name', ''),
                'usn': user.get('idno'),
                'department': user.get('department'),
                'year': user.get('year'),
                'section': user.get('section', 'A'),
                'linked_student_id': user.get('parent_id')
            }

# Password Reset Models
class ForgotPasswordRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str

# Email helper function
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import uuid

def send_reset_email(to_email: str, reset_token: str):
    smtp_email = os.environ.get('SMTP_EMAIL')
    smtp_password = os.environ.get('SMTP_PASSWORD')
    smtp_host = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
    smtp_port = int(os.environ.get('SMTP_PORT', 587))
    frontend_url = os.environ.get('FRONTEND_URL', 'http://localhost:3000')
    
    reset_link = f"{frontend_url}/reset-password?token={reset_token}"
    
    msg = MIMEMultipart()
    msg['From'] = smtp_email
    msg['To'] = to_email
    msg['Subject'] = 'JAIN LMS - Password Reset Request'
    
    body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; padding: 20px;">
        <h2 style="color: #1a365d;">Password Reset Request</h2>
        <p>You have requested to reset your password for your JAIN LMS account.</p>
        <p>Click the button below to reset your password:</p>
        <a href="{reset_link}" style="display: inline-block; padding: 12px 24px; background-color: #1a365d; color: white; text-decoration: none; border-radius: 5px; margin: 20px 0;">Reset Password</a>
        <p style="color: #666; font-size: 12px;">This link will expire in 15 minutes.</p>
        <p style="color: #666; font-size: 12px;">If you did not request this, please ignore this email.</p>
        <hr style="border: none; border-top: 1px solid #eee; margin: 20px 0;">
        <p style="color: #999; font-size: 11px;">JAIN University Learning Management System</p>
    </body>
    </html>
    """
    
    msg.attach(MIMEText(body, 'html'))
    
    try:
        server = smtplib.SMTP(smtp_host, smtp_port)
        server.starttls()
        server.login(smtp_email, smtp_password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        logging.error(f"Failed to send email: {e}")
        return False

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, email FROM users WHERE email = %s", (request.email,))
            user = cursor.fetchone()
            
            if not user:
                # Don't reveal if email exists
                return {"message": "If an account exists, a reset email has been sent."}
            
            # Generate reset token
            reset_token = str(uuid.uuid4())
            expires_at = datetime.now(timezone.utc) + timedelta(minutes=15)
            
            # Store token (using a simple approach - store in users table or create a new table)
            cursor.execute("""
                UPDATE users SET reset_token = %s, reset_token_expires = %s WHERE id = %s
            """, (reset_token, expires_at, user['id']))
            conn.commit()
            
            # Send email
            email_sent = send_reset_email(user['email'], reset_token)
            
            if not email_sent:
                raise HTTPException(status_code=500, detail="Failed to send reset email")
            
            return {"message": "If an account exists, a reset email has been sent."}

@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id FROM users WHERE reset_token = %s AND reset_token_expires > %s
            """, (request.token, datetime.now(timezone.utc)))
            user = cursor.fetchone()
            
            if not user:
                raise HTTPException(status_code=400, detail="Invalid or expired reset token")
            
            # Hash new password
            hashed = bcrypt.hashpw(request.new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # Update password and clear token
            cursor.execute("""
                UPDATE users SET password = %s, reset_token = NULL, reset_token_expires = NULL, 
                       must_change_password = FALSE WHERE id = %s
            """, (hashed, user['id']))
            conn.commit()
            
            return {"message": "Password reset successfully"}

@api_router.post("/auth/change-password")
async def change_password(request: ChangePasswordRequest, token: dict = Depends(verify_token)):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id, password FROM users WHERE id = %s", (token['user_id'],))
            user = cursor.fetchone()
            
            if not user:
                raise HTTPException(status_code=404, detail="User not found")
            
            # Verify current password
            if not bcrypt.checkpw(request.current_password.encode('utf-8'), user['password'].encode('utf-8')):
                raise HTTPException(status_code=400, detail="Current password is incorrect")
            
            # Hash new password
            hashed = bcrypt.hashpw(request.new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # Update password
            cursor.execute("""
                UPDATE users SET password = %s, must_change_password = FALSE WHERE id = %s
            """, (hashed, token['user_id']))
            conn.commit()
            
            return {"message": "Password changed successfully"}

# User Management
@api_router.get("/users")
async def get_users(token: dict = Depends(require_role('Admin'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT u.id, u.username, u.email, u.role, u.name as full_name, u.idno as usn, 
                       u.department, u.year, u.section, u.parent_id as linked_student_id, 
                       u.is_hod, u.hod_department, u.created_at,
                       ct.id as class_teacher_id
                FROM users u
                LEFT JOIN class_teachers ct ON u.id = ct.teacher_id
            """)
            return cursor.fetchall()

@api_router.get("/users/students")
async def get_students(
    department: Optional[str] = None, 
    year: Optional[str] = None,
    token: dict = Depends(require_role('Admin', 'Teacher'))
):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            query = """
                SELECT id, username, email, name as full_name, idno as usn, 
                       department, year, section, created_at 
                FROM users WHERE role = 'Student'
            """
            params = []
            
            if department and department != 'all':
                dept_ids = resolve_department_identifiers(department, cursor)
                query += " AND department IN %s"
                params.append(tuple(dept_ids))
            if year and year != 'all':
                query += " AND year = %s"
                params.append(year)
            
            query += " ORDER BY name"
            cursor.execute(query, params)
            return cursor.fetchall()


@api_router.get("/users/teachers")
async def get_teachers(
    department: Optional[str] = None,
    token: dict = Depends(require_role('Admin'))
):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            query = """
                SELECT id, username, email, name, name as full_name, 
                       department, is_hod, hod_department, created_at 
                FROM users WHERE role = 'Teacher'
            """
            params = []
            
            if department and department != 'all':
                dept_ids = resolve_department_identifiers(department, cursor)
                query += " AND department IN %s"
                params.append(tuple(dept_ids))
            
            query += " ORDER BY name"
            cursor.execute(query, params)
            return cursor.fetchall()


@api_router.get("/users/{user_id}")
async def get_user(user_id: int, token: dict = Depends(require_role('Admin', 'Teacher'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, username, email, role, name as full_name, idno as usn, 
                       department, year, section, parent_id as linked_student_id 
                FROM users WHERE id = %s
            """, (user_id,))
            user = cursor.fetchone()
            if not user:
                raise HTTPException(status_code=404, detail="User not found")
            return user

@api_router.post("/users")
async def create_user(user: UserCreate, token: dict = Depends(require_role('Admin'))):
    if not validate_domain(user.email, user.role):
        raise HTTPException(status_code=400, detail=f"{user.role}s must use @jainuniversity.ac.in domain")
    
    # Enforce Department, Year, Section for Students
    if user.role == 'Student':
        if not user.department or not user.year or not user.section:
            raise HTTPException(status_code=400, detail="Department, Year, and Section are required for Students")

    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Check if user already exists
            cursor.execute("SELECT id FROM users WHERE username = %s OR email = %s", (user.username, user.email))
            if cursor.fetchone():
                raise HTTPException(status_code=401, detail="Username or email already exists")
            
            # Hash password (default to 123456789 if not provided)
            password_to_hash = user.password if user.password else "123456789"
            hashed = bcrypt.hashpw(password_to_hash.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # Insert User
            cursor.execute("""
                INSERT INTO users (username, email, password, role, name, idno, department, year, section, parent_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (user.username, user.email, hashed, user.role, user.name, 
                  user.idno, user.department, user.year, user.section, user.parent_id))
            user_id = cursor.lastrowid
            
            # AUTOMATION: If Student, auto-create Parent Account
            if user.role == 'Student':
                # Example: john.doe -> john.doe22291
                usn_digits = "".join(filter(str.isdigit, user.idno))[-5:] if user.idno else "00000"
                parent_username = f"{user.username}{usn_digits}"
                parent_email = f"parent.{user.email}"
                
                # Check parent availability
                cursor.execute("SELECT id FROM users WHERE username = %s OR email = %s", (parent_username, parent_email))
                if not cursor.fetchone():
                    cursor.execute("""
                        INSERT INTO users (username, email, password, role, name, parent_id)
                        VALUES (%s, %s, %s, 'Parent', %s, %s)
                    """, (parent_username, parent_email, hashed, f"Parent of {user.name}", user_id))
            
            conn.commit()
            return {"id": user_id, "message": "User and linked accounts created successfully"}

# ==========================================
# DEPARTMENT MANAGEMENT
# ==========================================

@api_router.get("/departments")
async def get_departments(token: dict = Depends(verify_token)):
    """List all departments"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM departments ORDER BY name ASC")
            return cursor.fetchall()

@api_router.post("/departments")
async def add_department(dept: DepartmentCreate, token: dict = Depends(require_role('Admin'))):
    """Add a new department"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            try:
                cursor.execute("INSERT INTO departments (name, code) VALUES (%s, %s)", (dept.name, dept.code))
                conn.commit()
                return {"message": f"Department {dept.name} added"}
            except pymysql.err.IntegrityError:
                raise HTTPException(status_code=400, detail="Department or code already exists")

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: int, token: dict = Depends(require_role('Admin'))):
    """Delete a department"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM departments WHERE id = %s", (dept_id,))
            conn.commit()
            return {"message": "Department deleted"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: int, token: dict = Depends(require_role('Admin'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
            conn.commit()
            return {"message": "User deleted successfully"}

@api_router.post("/users/bulk-upload")
async def bulk_upload_students(
    file: UploadFile = File(...),
    department: Optional[str] = Form(None),
    year: Optional[str] = Form(None),
    section: Optional[str] = Form(None),
    token: dict = Depends(require_role('Admin'))
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    content = await file.read()
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    
    headers = [cell.value.lower().replace(' ', '_') if cell.value else '' for cell in sheet[1]]
    
    required = ['student_name', 'usn', 'department']
    missing = [r for r in required if r not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {missing}")
    
    students_created = 0
    parents_created = 0
    errors = []
    default_password = "123456789"
    hashed = bcrypt.hashpw(default_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = dict(zip(headers, row))
                
                if not row_dict.get('student_name') or not row_dict.get('usn'):
                    continue
                
                try:
                    usn = str(row_dict['usn']).upper()
                    usn_digits = ''.join(filter(str.isdigit, usn))[-5:]  # Last 5 digits
                    email = f"{usn.lower()}@jainuniversity.ac.in"
                    username = row_dict['student_name'].replace(' ', '.').lower()
                    
                    # Apply Overrides or fallbacks
                    final_dept = department if department else row_dict.get('department')
                    final_year = year if year else row_dict.get('year', '1')
                    final_section = section if section else row_dict.get('section', 'A')
                    
                    if not final_dept:
                        errors.append(f"Row {row_idx}: Missing department (no override provided)")
                        continue

                    # Create Student
                    cursor.execute("""
                        INSERT INTO users (username, email, password, role, name, idno, department, year, section)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (username, email, hashed, 'Student', 
                          row_dict['student_name'], usn, final_dept, final_year, final_section))
                    student_id = cursor.lastrowid
                    students_created += 1
                    
                    # Auto-create Parent Account
                    parent_username = f"{username}{usn_digits}"
                    parent_email = f"parent.{usn.lower()}@jainuniversity.ac.in"
                    
                    cursor.execute("""
                        INSERT INTO users (username, email, password, role, name, parent_id)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (parent_username, parent_email, hashed, 'Parent', 
                          f"Parent of {row_dict['student_name']}", student_id))
                    parents_created += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            conn.commit()
    
    return {
        "message": f"Successfully created {students_created} students and {parents_created} parent accounts",
        "students_created": students_created,
        "parents_created": parents_created,
        "errors": errors if errors else None
    }


@api_router.post("/users/link-parent")
async def link_parent_to_student(request: ParentLinkRequest, token: dict = Depends(require_role('Admin'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, idno FROM users 
                WHERE role = 'Student' 
                AND (username LIKE %s OR idno LIKE %s)
            """, (f"%{request.student_username}%", f"%{request.student_idno_digits}%"))
            student = cursor.fetchone()
            
            if not student:
                raise HTTPException(status_code=404, detail="Student not found")
            
            cursor.execute("""
                UPDATE users SET parent_id = %s WHERE id = %s
            """, (student['id'], request.parent_id))
            conn.commit()
            
            return {"message": "Parent linked to student successfully", "student_id": student['id']}

# Courses
@api_router.get("/courses")
async def get_courses(token: dict = Depends(verify_token)):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            if token['role'] == 'Teacher':
                cursor.execute("""
                    SELECT * FROM courses WHERE teacher_id = %s
                """, (token['user_id'],))
            else:
                cursor.execute("SELECT * FROM courses")
            return cursor.fetchall()

@api_router.post("/courses")
async def create_course(course: CourseCreate, token: dict = Depends(require_role('Admin', 'Teacher'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            teacher_id = token['user_id'] if token['role'] == 'Teacher' else None
            
            # Get teacher name
            teacher_name = None
            if teacher_id:
                cursor.execute("SELECT name FROM users WHERE id = %s", (teacher_id,))
                teacher = cursor.fetchone()
                teacher_name = teacher['name'] if teacher else None
            
            cursor.execute("""
                INSERT INTO courses (name, code, department, year, teacher_id, teacher_name)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (course.name, course.code, course.department, course.year, teacher_id, teacher_name))
            conn.commit()
            
            return {"id": cursor.lastrowid, "message": "Course created successfully"}

@api_router.get("/courses/{course_id}/students")
async def get_course_students(course_id: int, token: dict = Depends(require_role('Admin', 'Teacher'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Get course department and year
            cursor.execute("SELECT department, year FROM courses WHERE id = %s", (course_id,))
            course = cursor.fetchone()
            if not course:
                return []
            
            # Get students in same department/year
            cursor.execute("""
                SELECT id, username, email, name as full_name, idno as usn, department
                FROM users WHERE role = 'Student' AND department = %s
            """, (course.get('department'),))
            return cursor.fetchall()

# Grades
@api_router.get("/grades")
async def get_grades(token: dict = Depends(verify_token)):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            if token['role'] == 'Student':
                cursor.execute("""
                    SELECT g.id, g.student_id, g.course_id, g.course_name, 
                           g.title as assignment_name, g.marks as score, g.max_marks as max_score,
                           g.date as graded_at, 'Assignment' as grade_type
                    FROM grades g
                    WHERE g.student_id = %s
                    ORDER BY g.date DESC
                """, (token['user_id'],))
            elif token['role'] == 'Parent':
                cursor.execute("SELECT parent_id FROM users WHERE id = %s", (token['user_id'],))
                parent = cursor.fetchone()
                if not parent or not parent.get('parent_id'):
                    return []
                cursor.execute("""
                    SELECT g.id, g.student_id, g.course_id, g.course_name, 
                           g.title as assignment_name, g.marks as score, g.max_marks as max_score,
                           g.date as graded_at, u.name as student_name
                    FROM grades g
                    JOIN users u ON g.student_id = u.id
                    WHERE g.student_id = %s
                    ORDER BY g.date DESC
                """, (parent['parent_id'],))
            elif token['role'] == 'Teacher':
                cursor.execute("""
                    SELECT g.id, g.student_id, g.course_id, g.course_name, 
                           g.title as assignment_name, g.marks as score, g.max_marks as max_score,
                           g.date as graded_at, u.name as student_name, u.idno as usn
                    FROM grades g
                    JOIN users u ON g.student_id = u.id
                    WHERE g.graded_by = %s
                    ORDER BY g.date DESC
                """, (token['user_id'],))
            else:
                cursor.execute("""
                    SELECT g.id, g.student_id, g.course_id, g.course_name, 
                           g.title as assignment_name, g.marks as score, g.max_marks as max_score,
                           g.date as graded_at, u.name as student_name
                    FROM grades g
                    JOIN users u ON g.student_id = u.id
                    ORDER BY g.date DESC
                """)
            return cursor.fetchall()

@api_router.post("/grades")
async def create_grade(grade: GradeCreate, token: dict = Depends(require_role('Admin', 'Teacher'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                INSERT INTO grades (student_id, course_id, course_name, title, marks, max_marks, graded_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (grade.student_id, grade.course_id, grade.course_name, grade.title,
                  grade.marks, grade.max_marks, token['user_id']))
            conn.commit()
            return {"id": cursor.lastrowid, "message": "Grade posted successfully"}

# Attendance
@api_router.get("/attendance")
async def get_attendance(token: dict = Depends(verify_token)):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            if token['role'] == 'Student':
                cursor.execute("""
                    SELECT a.id, a.course_id, a.course_name, a.date, a.records
                    FROM attendance a
                    ORDER BY a.date DESC
                """)
                all_attendance = cursor.fetchall()
                # Filter for this student from JSON records
                student_attendance = []
                for record in all_attendance:
                    records = record.get('records')
                    if isinstance(records, str):
                        records = json.loads(records)
                    if records:
                        for r in records:
                            if r.get('student_id') == token['user_id']:
                                student_attendance.append({
                                    'id': record['id'],
                                    'course_id': record['course_id'],
                                    'course_name': record['course_name'],
                                    'date': record['date'],
                                    'status': r.get('status', 'Present')
                                })
                return student_attendance
            elif token['role'] == 'Parent':
                cursor.execute("SELECT parent_id FROM users WHERE id = %s", (token['user_id'],))
                parent = cursor.fetchone()
                if not parent or not parent.get('parent_id'):
                    return []
                student_id = parent['parent_id']
                cursor.execute("SELECT name FROM users WHERE id = %s", (student_id,))
                student = cursor.fetchone()
                student_name = student['name'] if student else 'Student'
                
                cursor.execute("""
                    SELECT a.id, a.course_id, a.course_name, a.date, a.records
                    FROM attendance a
                    ORDER BY a.date DESC
                """)
                all_attendance = cursor.fetchall()
                student_attendance = []
                for record in all_attendance:
                    records = record.get('records')
                    if isinstance(records, str):
                        records = json.loads(records)
                    if records:
                        for r in records:
                            if r.get('student_id') == student_id:
                                student_attendance.append({
                                    'id': record['id'],
                                    'course_id': record['course_id'],
                                    'course_name': record['course_name'],
                                    'date': record['date'],
                                    'status': r.get('status', 'Present'),
                                    'student_name': student_name
                                })
                return student_attendance
            elif token['role'] == 'Teacher':
                cursor.execute("""
                    SELECT a.*, c.name as course_display_name
                    FROM attendance a
                    LEFT JOIN courses c ON a.course_id = c.id
                    WHERE a.taken_by = %s
                    ORDER BY a.date DESC
                """, (token['user_id'],))
                return cursor.fetchall()
            else:
                cursor.execute("""
                    SELECT a.*, c.name as course_display_name
                    FROM attendance a
                    LEFT JOIN courses c ON a.course_id = c.id
                    ORDER BY a.date DESC
                """)
                return cursor.fetchall()

@api_router.post("/attendance")
async def mark_attendance(attendance: AttendanceCreate, token: dict = Depends(require_role('Admin', 'Teacher'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            records_json = json.dumps([{"student_id": r.student_id, "status": r.status} for r in attendance.records])
            
            cursor.execute("""
                INSERT INTO attendance (course_id, course_name, department, year, date, records, taken_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (attendance.course_id, attendance.course_name, attendance.department, 
                  attendance.year, attendance.date, records_json, token['user_id']))
            conn.commit()
            return {"message": "Attendance marked successfully"}

@api_router.post("/attendance/bulk")
async def bulk_mark_attendance(attendance: AttendanceCreate, token: dict = Depends(require_role('Admin', 'Teacher'))):
    return await mark_attendance(attendance, token)

# Classwork
@api_router.get("/classwork")
async def get_classwork(token: dict = Depends(verify_token)):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            if token['role'] == 'Student':
                cursor.execute("SELECT department FROM users WHERE id = %s", (token['user_id'],))
                user = cursor.fetchone()
                dept = user.get('department') if user else None
                
                cursor.execute("""
                    SELECT cw.*, c.name as course_name
                    FROM classwork cw
                    LEFT JOIN courses c ON cw.course_id = c.id
                    WHERE cw.department = %s OR cw.department IS NULL
                    ORDER BY cw.created_at DESC
                """, (dept,))
            elif token['role'] == 'Teacher':
                cursor.execute("""
                    SELECT cw.*, c.name as course_name
                    FROM classwork cw
                    LEFT JOIN courses c ON cw.course_id = c.id
                    WHERE cw.uploaded_by = %s
                    ORDER BY cw.created_at DESC
                """, (token['user_id'],))
            else:
                cursor.execute("""
                    SELECT cw.*, c.name as course_name
                    FROM classwork cw
                    LEFT JOIN courses c ON cw.course_id = c.id
                    ORDER BY cw.created_at DESC
                """)
            return cursor.fetchall()

@api_router.post("/classwork")
async def create_classwork(classwork: ClassworkCreate, token: dict = Depends(require_role('Admin', 'Teacher'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                INSERT INTO classwork (course_id, department, year, type, title, description, due_date, max_marks, uploaded_by)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (classwork.course_id, classwork.department, classwork.year, classwork.type,
                  classwork.title, classwork.description, classwork.due_date, classwork.max_marks, token['user_id']))
            conn.commit()
            return {"id": cursor.lastrowid, "message": "Classwork created successfully"}

# Submissions
@api_router.get("/submissions")
async def get_submissions(token: dict = Depends(verify_token)):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            if token['role'] == 'Student':
                cursor.execute("""
                    SELECT s.*, cw.title as classwork_title, c.name as course_name
                    FROM submissions s
                    JOIN classwork cw ON s.classwork_id = cw.id
                    LEFT JOIN courses c ON cw.course_id = c.id
                    WHERE s.student_id = %s
                    ORDER BY s.submitted_at DESC
                """, (token['user_id'],))
            elif token['role'] == 'Teacher':
                cursor.execute("""
                    SELECT s.*, cw.title as classwork_title, u.idno as usn
                    FROM submissions s
                    JOIN classwork cw ON s.classwork_id = cw.id
                    JOIN users u ON s.student_id = u.id
                    WHERE cw.uploaded_by = %s
                    ORDER BY s.submitted_at DESC
                """, (token['user_id'],))
            else:
                cursor.execute("""
                    SELECT s.*, cw.title as classwork_title
                    FROM submissions s
                    JOIN classwork cw ON s.classwork_id = cw.id
                    ORDER BY s.submitted_at DESC
                """)
            return cursor.fetchall()

@api_router.post("/submissions")
async def create_submission(submission: SubmissionCreate, token: dict = Depends(require_role('Student'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Get student name
            cursor.execute("SELECT name FROM users WHERE id = %s", (token['user_id'],))
            user = cursor.fetchone()
            student_name = user['name'] if user else 'Student'
            
            cursor.execute("""
                INSERT INTO submissions (classwork_id, student_id, student_name, content, status)
                VALUES (%s, %s, %s, %s, 'Submitted')
                ON DUPLICATE KEY UPDATE content = %s, submitted_at = CURRENT_TIMESTAMP
            """, (submission.classwork_id, token['user_id'], student_name,
                  submission.content, submission.content))
            conn.commit()
            return {"message": "Submission created successfully"}

# Dashboard Stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(token: dict = Depends(verify_token)):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            stats = {}
            
            if token['role'] == 'Admin':
                cursor.execute("SELECT COUNT(*) as count FROM users WHERE role = 'Student'")
                stats['total_students'] = cursor.fetchone()['count']
                cursor.execute("SELECT COUNT(*) as count FROM users WHERE role = 'Teacher'")
                stats['total_teachers'] = cursor.fetchone()['count']
                cursor.execute("SELECT COUNT(*) as count FROM courses")
                stats['total_courses'] = cursor.fetchone()['count']
                cursor.execute("SELECT COUNT(*) as count FROM users WHERE role = 'Parent'")
                stats['total_parents'] = cursor.fetchone()['count']
            
            elif token['role'] == 'Teacher':
                cursor.execute("SELECT COUNT(*) as count FROM courses WHERE teacher_id = %s", (token['user_id'],))
                stats['my_courses'] = cursor.fetchone()['count']
                cursor.execute("SELECT COUNT(*) as count FROM users WHERE role = 'Student'")
                stats['my_students'] = cursor.fetchone()['count']
                cursor.execute("SELECT COUNT(*) as count FROM classwork WHERE uploaded_by = %s", (token['user_id'],))
                stats['total_classwork'] = cursor.fetchone()['count']
            
            elif token['role'] == 'Student':
                cursor.execute("SELECT COUNT(*) as count FROM courses")
                stats['enrolled_courses'] = cursor.fetchone()['count']
                cursor.execute("""
                    SELECT AVG(marks/max_marks * 100) as avg FROM grades WHERE student_id = %s AND max_marks > 0
                """, (token['user_id'],))
                result = cursor.fetchone()
                avg = result['avg'] if result and result['avg'] else 0
                stats['average_grade'] = round(float(avg), 1) if avg else 0
                stats['attendance_rate'] = 100  # Default, would need to calculate from JSON
            
            elif token['role'] == 'Parent':
                cursor.execute("SELECT parent_id FROM users WHERE id = %s", (token['user_id'],))
                parent = cursor.fetchone()
                if parent and parent.get('parent_id'):
                    student_id = parent['parent_id']
                    cursor.execute("SELECT name FROM users WHERE id = %s", (student_id,))
                    student = cursor.fetchone()
                    stats['student_name'] = student['name'] if student else 'Not linked'
                    cursor.execute("""
                        SELECT AVG(marks/max_marks * 100) as avg FROM grades WHERE student_id = %s AND max_marks > 0
                    """, (student_id,))
                    result = cursor.fetchone()
                    avg = result['avg'] if result and result['avg'] else 0
                    stats['student_average'] = round(float(avg), 1) if avg else 0
                    stats['student_attendance'] = 100
                else:
                    stats['student_name'] = 'Not linked'
                    stats['student_average'] = 0
                    stats['student_attendance'] = 100
            
            return stats

# Students list for dropdowns
@api_router.get("/students")
async def get_students(token: dict = Depends(require_role('Admin', 'Teacher'))):
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, username, email, name as full_name, idno as usn, department, section 
                FROM users WHERE role = 'Student'
            """)
            return cursor.fetchall()

# ==========================================
# TIMETABLE GENERATION
# ==========================================

# Time slots configuration
TIME_SLOTS = [
    {"slot": 1, "start": "08:45", "end": "09:45", "type": "class"},
    {"slot": 2, "start": "09:45", "end": "10:45", "type": "class"},
    {"slot": 3, "start": "11:00", "end": "12:00", "type": "class"},  # After break
    {"slot": 4, "start": "12:50", "end": "13:50", "type": "class"},  # After lunch
    {"slot": 5, "start": "13:50", "end": "14:50", "type": "class"},
    {"slot": 6, "start": "14:50", "end": "15:50", "type": "class"},
]

DAYS_OF_WEEK = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

class TimetableSlotCreate(BaseModel):
    department: str
    year: str
    section: str
    day_of_week: str
    slot_number: int
    course_id: int
    teacher_id: int
    room: Optional[str] = None

class ClassTeacherAssign(BaseModel):
    department: str
    year: str
    section: str
    teacher_id: int

@api_router.get("/timetable/slots-config")
async def get_slots_config(token: dict = Depends(verify_token)):
    """Get the time slot configuration"""
    return {"slots": TIME_SLOTS, "days": DAYS_OF_WEEK}

@api_router.get("/timetable")
async def get_timetable(
    department: Optional[str] = None,
    year: Optional[str] = None,
    section: Optional[str] = None,
    token: dict = Depends(verify_token)
):
    """Get timetable for a specific class or all classes"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            query = """
                SELECT ts.*, c.name as course_name, c.code as course_code,
                       u.name as teacher_name
                FROM timetable_slots ts
                LEFT JOIN courses c ON ts.course_id = c.id
                LEFT JOIN users u ON ts.teacher_id = u.id
                WHERE 1=1
            """
            params = []
            
            if department:
                query += " AND ts.department = %s"
                params.append(department)
            if year:
                query += " AND ts.year = %s"
                params.append(year)
            if section:
                query += " AND ts.section = %s"
                params.append(section)
            
            query += " ORDER BY ts.day_of_week, ts.slot_number"
            cursor.execute(query, params)
            slots = cursor.fetchall()
            
            # Also get class teacher if filtering by specific class
            class_teacher = None
            if department and year and section:
                cursor.execute("""
                    SELECT ct.*, u.name as teacher_name 
                    FROM class_teachers ct
                    JOIN users u ON ct.teacher_id = u.id
                    WHERE ct.department = %s AND ct.year = %s AND ct.section = %s
                """, (department, year, section))
                class_teacher = cursor.fetchone()
            
            return {"slots": slots, "class_teacher": class_teacher, "config": TIME_SLOTS}

@api_router.get("/timetable/today")
async def get_today_timetable(token: dict = Depends(verify_token)):
    """Get today's timetable for the logged-in student"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Get user details
            cursor.execute("SELECT department, year, section FROM users WHERE id = %s", (token['user_id'],))
            user = cursor.fetchone()
            
            if not user or not user.get('department') or not user.get('year'):
                return {"slots": [], "message": "User profile incomplete"}
            
            # Get current day
            from datetime import datetime
            today = datetime.now().strftime("%A")
            
            section = user.get('section', 'A')
            
            cursor.execute("""
                SELECT ts.*, c.name as course_name, c.code as course_code,
                       u.name as teacher_name
                FROM timetable_slots ts
                LEFT JOIN courses c ON ts.course_id = c.id
                LEFT JOIN users u ON ts.teacher_id = u.id
                WHERE ts.department = %s AND ts.year = %s AND ts.section = %s AND ts.day_of_week = %s
                ORDER BY ts.slot_number
            """, (user['department'], user['year'], section, today))
            slots = cursor.fetchall()
            
            # Mark current and next class
            current_time = datetime.now().strftime("%H:%M")
            for i, slot in enumerate(slots):
                slot_config = next((s for s in TIME_SLOTS if s['slot'] == slot['slot_number']), None)
                if slot_config:
                    slot['start_time'] = slot_config['start']
                    slot['end_time'] = slot_config['end']
                    slot['is_current'] = slot_config['start'] <= current_time <= slot_config['end']
                    slot['is_next'] = i > 0 and slots[i-1].get('is_current', False) == False and current_time < slot_config['start']
            
            return {"slots": slots, "today": today, "current_time": current_time}

@api_router.post("/timetable/slot")
async def create_timetable_slot(
    slot: TimetableSlotCreate,
    token: dict = Depends(require_role('Admin'))
):
    """Create or update a single timetable slot"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Check for teacher collision (same teacher, same day, same slot)
            cursor.execute("""
                SELECT id FROM timetable_slots 
                WHERE teacher_id = %s AND day_of_week = %s AND slot_number = %s
                AND NOT (department = %s AND year = %s AND section = %s)
            """, (slot.teacher_id, slot.day_of_week, slot.slot_number,
                  slot.department, slot.year, slot.section))
            
            if cursor.fetchone():
                raise HTTPException(status_code=400, 
                    detail=f"Teacher already has a class on {slot.day_of_week} at slot {slot.slot_number}")
            
            # Check teacher's daily limit (max 2 classes per day)
            cursor.execute("""
                SELECT COUNT(*) as count FROM timetable_slots 
                WHERE teacher_id = %s AND day_of_week = %s
                AND NOT (department = %s AND year = %s AND section = %s AND slot_number = %s)
            """, (slot.teacher_id, slot.day_of_week, 
                  slot.department, slot.year, slot.section, slot.slot_number))
            count = cursor.fetchone()['count']
            
            if count >= 2:
                raise HTTPException(status_code=400, 
                    detail=f"Teacher already has maximum 2 classes on {slot.day_of_week}")
            
            # Insert or update
            cursor.execute("""
                INSERT INTO timetable_slots 
                (department, year, section, day_of_week, slot_number, course_id, teacher_id, room)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE 
                course_id = VALUES(course_id), teacher_id = VALUES(teacher_id), room = VALUES(room)
            """, (slot.department, slot.year, slot.section, slot.day_of_week,
                  slot.slot_number, slot.course_id, slot.teacher_id, slot.room))
            conn.commit()
            return {"message": "Slot updated successfully"}

@api_router.get("/users/teachers/subjects")
async def get_teachers_with_subjects(department: str, token: dict = Depends(require_role('Admin'))):
    """Fetch teachers and the courses common in their department"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Resolve department name if code is provided
            cursor.execute("SELECT name FROM departments WHERE code = %s OR name = %s", (department, department))
            dept_res = cursor.fetchone()
            dept_name = dept_res['name'] if dept_res else department

            # Get teachers
            cursor.execute("SELECT id, name FROM users WHERE role = 'Teacher'")
            teachers = cursor.fetchall()
            
            # Get courses for this department (match by code or name)
            cursor.execute("""
                SELECT id, name, code FROM courses 
                WHERE department = %s OR department = %s
            """, (department, dept_name))
            courses = cursor.fetchall()
            
            return {"teachers": teachers, "courses": courses}

@api_router.post("/timetable/generate")
async def generate_timetable(req: TimetableGenerateRequest, token: dict = Depends(require_role('Admin'))):
    """
    Generate a collision-free timetable.
    Algorithm:
    1. Collect all requirements (total slots needed per teacher/course).
    2. Iterate through all available days and slots for this section.
    3. For each slot, find a teacher from the pairings who is:
       - Available (not teaching elsewhere in the same slot).
       - Hasn't hit their daily limit of 2 classes.
       - Still needs more slots to meet their weekly target.
    """
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    all_slots = range(1, 9) # 8 slots per day
    
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # 1. Clear existing timetable for this specific section
            cursor.execute("""
                DELETE FROM timetable_slots 
                WHERE department = %s AND year = %s AND section = %s
            """, (req.department, req.year, req.section))
            
            # 2. Track assignments for THIS generation
            # requirements: {(course_id, teacher_id): slots_remaining}
            requirements = {(p.course_id, p.teacher_id): p.slots_per_week for p in req.pairings}
            
            # 3. Greedy assignment with collision checks
            for day in days:
                for slot_num in all_slots:
                    # Randomize pairings order to avoid bias
                    pairing_items = list(requirements.items())
                    random.shuffle(pairing_items)
                    
                    for (course_id, teacher_id), remaining in pairing_items:
                        if remaining <= 0:
                            continue
                            
                        # Check teacher availability (global collision)
                        cursor.execute("""
                            SELECT id FROM timetable_slots 
                            WHERE teacher_id = %s AND day_of_week = %s AND slot_number = %s
                        """, (teacher_id, day, slot_num))
                        if cursor.fetchone():
                            continue # Teacher is busy elsewhere
                            
                        # Check teacher daily limit (max 2)
                        cursor.execute("""
                            SELECT COUNT(*) as count FROM timetable_slots 
                            WHERE teacher_id = %s AND day_of_week = %s
                        """, (teacher_id, day))
                        daily_count = cursor.fetchone()['count']
                        if daily_count >= 2:
                            continue
                            
                        # Assign slot
                        cursor.execute("""
                            INSERT INTO timetable_slots 
                            (department, year, section, day_of_week, slot_number, course_id, teacher_id)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (req.department, req.year, req.section, day, slot_num, course_id, teacher_id))
                        
                        requirements[(course_id, teacher_id)] -= 1
                        break # Found a teacher for this slot, move to next slot
            
            conn.commit()
            
            # Check for unassigned slots
            total_remaining = sum(requirements.values())
            if total_remaining > 0:
                return {
                    "message": f"Timetable generated with {total_remaining} unassigned slots due to collisions.",
                    "status": "partial"
                }
                
            return {"message": "Timetable generated successfully!", "status": "success"}
            conn.commit()
            
            return {"message": "Slot created/updated successfully"}

@api_router.delete("/timetable/slot/{slot_id}")
async def delete_timetable_slot(
    slot_id: int,
    token: dict = Depends(require_role('Admin'))
):
    """Delete a timetable slot"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM timetable_slots WHERE id = %s", (slot_id,))
            conn.commit()
            return {"message": "Slot deleted successfully"}

@api_router.post("/timetable/class-teacher")
async def assign_class_teacher(
    assignment: ClassTeacherAssign,
    token: dict = Depends(require_role('Admin'))
):
    """Assign a class teacher to a section"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Check if teacher is an HOD
            cursor.execute("SELECT is_hod FROM users WHERE id = %s", (assignment.teacher_id,))
            user = cursor.fetchone()
            if user and user['is_hod']:
                raise HTTPException(status_code=400, detail="This teacher is an HOD and cannot be a Class Teacher.")

            cursor.execute("""
                INSERT INTO class_teachers (department, year, section, teacher_id)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE teacher_id = VALUES(teacher_id)
            """, (assignment.department, assignment.year, assignment.section, assignment.teacher_id))
            conn.commit()
            return {"message": "Class teacher assigned successfully"}

@api_router.get("/timetable/class-teachers")
async def get_class_teachers(
    department: Optional[str] = None,
    token: dict = Depends(require_role('Admin', 'Teacher'))
):
    """Get all class teacher assignments"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            query = """
                SELECT ct.*, u.name as teacher_name, u.email as teacher_email
                FROM class_teachers ct
                JOIN users u ON ct.teacher_id = u.id
            """
            params = []
            if department:
                query += " WHERE ct.department = %s"
                params.append(department)
            
            query += " ORDER BY ct.department, ct.year, ct.section"
            cursor.execute(query, params)
            return cursor.fetchall()

@api_router.get("/timetable/class-teacher/check")
async def check_class_teacher(
    department: str,
    year: str,
    section: str,
    token: dict = Depends(require_role('Admin'))
):
    """Check if a class teacher already exists for a section"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT ct.*, u.name as teacher_name 
                FROM class_teachers ct
                JOIN users u ON ct.teacher_id = u.id
                WHERE ct.department = %s AND ct.year = %s AND ct.section = %s
            """, (department, year, section))
            return cursor.fetchone()


# ==========================================
# LEAVE REQUEST SYSTEM
# ==========================================

class LeaveRequestCreate(BaseModel):
    leave_type: str  # sick, personal, emergency
    start_date: str
    end_date: str
    reason: str

class LeaveApproval(BaseModel):
    status: str  # approved, rejected
    remarks: Optional[str] = None

@api_router.post("/leave/request")
async def create_leave_request(
    request: LeaveRequestCreate,
    token: dict = Depends(verify_token)
):
    """Student submits a leave request"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Get student info
            cursor.execute("SELECT id, name, department, year, section FROM users WHERE id = %s", (token['user_id'],))
            student = cursor.fetchone()
            
            if not student:
                raise HTTPException(status_code=404, detail="User not found")
            
            # Find class teacher for this student's specific class (Dept + Year + Section)
            cursor.execute("""
                SELECT ct.teacher_id, u.name as teacher_name
                FROM class_teachers ct
                JOIN users u ON ct.teacher_id = u.id
                WHERE ct.department = %s AND ct.year = %s AND ct.section = %s
            """, (student['department'], student['year'], student['section']))
            class_teacher = cursor.fetchone()
            
            cursor.execute("""
                INSERT INTO leave_requests 
                (student_id, student_name, department, year, section, leave_type, start_date, end_date, reason, 
                 status, class_teacher_id, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending', %s, NOW())
            """, (token['user_id'], student['name'], student['department'], student['year'], student['section'],
                  request.leave_type, request.start_date, request.end_date, request.reason,
                  class_teacher['teacher_id'] if class_teacher else None))
            conn.commit()
            
            return {"message": "Leave request submitted successfully"}

@api_router.get("/leave/my-requests")
async def get_my_leave_requests(token: dict = Depends(verify_token)):
    """Get leave requests for the logged-in student"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM leave_requests WHERE student_id = %s ORDER BY created_at DESC
            """, (token['user_id'],))
            return cursor.fetchall()

@api_router.get("/leave/requests")
async def get_leave_requests_for_teacher(
    status: Optional[str] = None,
    token: dict = Depends(require_role('Teacher'))
):
    """Class teacher gets leave requests for their class"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            query = """
                SELECT lr.*, u.name as student_name, u.email as student_email
                FROM leave_requests lr
                JOIN users u ON lr.student_id = u.id
                WHERE lr.class_teacher_id = %s
            """
            params = [token['user_id']]
            
            if status:
                query += " AND lr.status = %s"
                params.append(status)
            
            query += " ORDER BY lr.created_at DESC"
            cursor.execute(query, params)
            return cursor.fetchall()

@api_router.get("/leave/hod-requests")
async def get_leave_requests_for_hod(
    status: Optional[str] = None,
    token: dict = Depends(verify_token)
):
    """HOD gets forwarded leave requests"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Verify user is HOD
            cursor.execute("SELECT is_hod, hod_department FROM users WHERE id = %s", (token['user_id'],))
            user = cursor.fetchone()
            
            if not user or not user.get('is_hod'):
                raise HTTPException(status_code=403, detail="Only HODs can access this")
            
            dept_ids = resolve_department_identifiers(user['hod_department'], cursor)
            query = """
                SELECT lr.*, u.name as student_name, t.name as teacher_name
                FROM leave_requests lr
                JOIN users u ON lr.student_id = u.id
                LEFT JOIN users t ON lr.class_teacher_id = t.id
                WHERE lr.department IN %s AND lr.status = 'forwarded_to_hod'
            """
            params = [tuple(dept_ids)]

            
            if status and status != 'forwarded_to_hod':
                query += " AND lr.status = %s"
                params.append(status)
            
            query += " ORDER BY lr.created_at DESC"
            cursor.execute(query, params)
            return cursor.fetchall()

@api_router.put("/leave/{request_id}/approve")
async def approve_leave_request(
    request_id: int,
    approval: LeaveApproval,
    token: dict = Depends(require_role('Teacher'))
):
    """Class teacher approves/rejects a leave request"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Verify this teacher owns this request
            cursor.execute("""
                SELECT * FROM leave_requests WHERE id = %s AND class_teacher_id = %s
            """, (request_id, token['user_id']))
            leave_req = cursor.fetchone()
            
            if not leave_req:
                raise HTTPException(status_code=404, detail="Leave request not found")
            
            # Get teacher name for signature
            cursor.execute("SELECT name FROM users WHERE id = %s", (token['user_id'],))
            teacher = cursor.fetchone()
            
            cursor.execute("""
                UPDATE leave_requests 
                SET status = %s, teacher_remarks = %s, approved_by = %s, approved_at = NOW()
                WHERE id = %s
            """, (approval.status, approval.remarks, teacher['name'], request_id))
            conn.commit()
            
            return {"message": f"Leave request {approval.status}"}

@api_router.put("/leave/{request_id}/forward-to-hod")
async def forward_to_hod(
    request_id: int,
    token: dict = Depends(require_role('Teacher'))
):
    """Class teacher forwards leave request to HOD for approval"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                UPDATE leave_requests SET status = 'forwarded_to_hod' WHERE id = %s AND class_teacher_id = %s
            """, (request_id, token['user_id']))
            conn.commit()
            return {"message": "Leave request forwarded to HOD"}

@api_router.put("/leave/{request_id}/hod-approve")
async def hod_approve_leave(
    request_id: int,
    approval: LeaveApproval,
    token: dict = Depends(verify_token)
):
    """HOD approves/rejects a forwarded leave request"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Verify user is HOD
            cursor.execute("SELECT is_hod, hod_department, name FROM users WHERE id = %s", (token['user_id'],))
            user = cursor.fetchone()
            
            if not user or not user.get('is_hod'):
                raise HTTPException(status_code=403, detail="Only HODs can approve")
            
            dept_ids = resolve_department_identifiers(user['hod_department'], cursor)
            cursor.execute("""
                UPDATE leave_requests 
                SET status = %s, hod_remarks = %s, hod_approved_by = %s, hod_approved_at = NOW()
                WHERE id = %s AND department IN %s
            """, (f"hod_{approval.status}", approval.remarks, user['name'], request_id, tuple(dept_ids)))

            conn.commit()
            
            return {"message": f"Leave request {approval.status} by HOD"}

@api_router.get("/leave/{request_id}/pdf")
async def get_leave_pdf_data(
    request_id: int,
    token: dict = Depends(verify_token)
):
    """Get leave request data for PDF generation (generated client-side)"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT lr.*, 
                       u.name as student_name, u.email as student_email, u.idno as usn,
                       t.name as teacher_name, t.email as teacher_email
                FROM leave_requests lr
                JOIN users u ON lr.student_id = u.id
                LEFT JOIN users t ON lr.class_teacher_id = t.id
                WHERE lr.id = %s AND (lr.student_id = %s OR lr.class_teacher_id = %s)
            """, (request_id, token['user_id'], token['user_id']))
            leave_req = cursor.fetchone()
            
            if not leave_req:
                raise HTTPException(status_code=404, detail="Leave request not found")
            
            if leave_req['status'] not in ['approved', 'hod_approved']:
                raise HTTPException(status_code=400, detail="Leave request not yet approved")
            
            return {
                "student_name": leave_req['student_name'],
                "usn": leave_req['usn'],
                "department": leave_req['department'],
                "year": leave_req['year'],
                "leave_type": leave_req['leave_type'],
                "start_date": str(leave_req['start_date']),
                "end_date": str(leave_req['end_date']),
                "reason": leave_req['reason'],
                "approved_by": leave_req.get('approved_by') or leave_req.get('hod_approved_by'),
                "approved_at": str(leave_req.get('approved_at') or leave_req.get('hod_approved_at')),
                "remarks": leave_req.get('teacher_remarks') or leave_req.get('hod_remarks')
            }

# ==========================================
# HOD ROLE & DEPARTMENT OVERSIGHT
# ==========================================

class HODAssignment(BaseModel):
    teacher_id: int
    department: str

class SummonStudent(BaseModel):
    student_id: int
    reason: str
    scheduled_time: Optional[str] = None

class AttendanceSessionStart(BaseModel):
    course_id: int
    lat: float
    lng: float
    radius: int = 20

class AttendanceMark(BaseModel):
    otp: str
    lat: float
    lng: float

class ManualAttendance(BaseModel):
    session_id: int
    student_id: int
    status: str = 'present'

@api_router.post("/hod/assign")
async def assign_hod(
    assignment: HODAssignment,
    token: dict = Depends(require_role('Admin'))
):
    """Admin assigns HOD role to a teacher"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Check if teacher is already a class teacher
            cursor.execute("SELECT id FROM class_teachers WHERE teacher_id = %s", (assignment.teacher_id,))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="This teacher is already a Class Teacher. Remove that role first.")

            # Remove existing HOD for this department
            cursor.execute("""
                UPDATE users SET is_hod = FALSE, hod_department = NULL 
                WHERE hod_department = %s
            """, (assignment.department,))
            
            # Assign new HOD
            cursor.execute("""
                UPDATE users SET is_hod = TRUE, hod_department = %s 
                WHERE id = %s AND role = 'Teacher'
            """, (assignment.department, assignment.teacher_id))
            conn.commit()
            
            return {"message": f"HOD assigned for {assignment.department}"}

@api_router.get("/hod/list")
async def get_all_hods(token: dict = Depends(require_role('Admin'))):
    """Get list of all HODs"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, name, email, hod_department 
                FROM users WHERE is_hod = TRUE
            """)
            return cursor.fetchall()

@api_router.get("/hod/check")
async def check_hod(
    department: str,
    token: dict = Depends(require_role('Admin'))
):
    """Check if an HOD already exists for a department"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, name, email 
                FROM users 
                WHERE is_hod = TRUE AND hod_department = %s
            """, (department,))
            return cursor.fetchone()


@api_router.get("/hod/department-overview")
async def get_department_overview(token: dict = Depends(verify_token)):
    """HOD gets overview of their department"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Verify user is HOD
            cursor.execute("SELECT is_hod, hod_department FROM users WHERE id = %s", (token['user_id'],))
            user = cursor.fetchone()
            
            if not user or not user.get('is_hod'):
                raise HTTPException(status_code=403, detail="Only HODs can access this")
            
            dept = user['hod_department']
            dept_ids = resolve_department_identifiers(dept, cursor)
            dept_tuple = tuple(dept_ids)
            
            # Resolve full department name for display
            display_name = dept
            cursor.execute("SELECT name FROM departments WHERE code = %s OR name = %s", (dept, dept))
            dept_res = cursor.fetchone()
            if dept_res:
                display_name = dept_res['name']

            # Get counts
            cursor.execute("""
                SELECT COUNT(*) as count FROM users 
                WHERE department IN %s AND role = 'Student'
            """, (dept_tuple,))
            student_count = cursor.fetchone()['count']
            
            cursor.execute("""
                SELECT COUNT(*) as count FROM users 
                WHERE department IN %s AND role = 'Teacher'
            """, (dept_tuple,))
            teacher_count = cursor.fetchone()['count']
            
            cursor.execute("""
                SELECT COUNT(*) as count FROM courses 
                WHERE department IN %s
            """, (dept_tuple,))
            course_count = cursor.fetchone()['count']
            
            # Get pending leave requests
            cursor.execute("""
                SELECT COUNT(*) as count FROM leave_requests 
                WHERE department IN %s AND status = 'forwarded_to_hod'
            """, (dept_tuple,))
            pending_leaves = cursor.fetchone()['count']
            
            # Get teachers list
            cursor.execute("""
                SELECT id, name, email FROM users 
                WHERE department IN %s AND role = 'Teacher'
            """, (dept_tuple,))
            teachers = cursor.fetchall()
            
            return {
                "department": display_name,
                "student_count": student_count,
                "teacher_count": teacher_count,
                "course_count": course_count,
                "pending_leaves": pending_leaves,
                "teachers": teachers
            }



@api_router.get("/hod/teacher/{teacher_id}/students")
async def get_teacher_students_performance(
    teacher_id: int,
    token: dict = Depends(verify_token)
):
    """HOD views students taught by a teacher, ranked by performance"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Verify user is HOD
            cursor.execute("SELECT is_hod, hod_department FROM users WHERE id = %s", (token['user_id'],))
            user = cursor.fetchone()
            
            if not user or not user.get('is_hod'):
                raise HTTPException(status_code=403, detail="Only HODs can access this")
            
            dept_ids = resolve_department_identifiers(user['hod_department'], cursor)
            dept_tuple = tuple(dept_ids)

            # Get teacher's courses
            cursor.execute("""
                SELECT DISTINCT c.id, c.name, c.code
                FROM courses c
                WHERE c.teacher_id = %s AND c.department IN %s
            """, (teacher_id, dept_tuple))
            courses = cursor.fetchall()
            
            # For each course, get students with grades
            result = []
            for course in courses:
                cursor.execute("""
                    SELECT u.id, u.name, u.idno as usn, 
                           COALESCE(AVG(g.marks), 0) as average_marks
                    FROM users u
                    LEFT JOIN grades g ON u.id = g.student_id AND g.course_id = %s
                    WHERE u.department IN %s AND u.role = 'Student'
                    GROUP BY u.id, u.name, u.idno
                    ORDER BY average_marks DESC
                """, (course['id'], dept_tuple))
                students = cursor.fetchall()

                
                result.append({
                    "course": course,
                    "students": students
                })
            
            return result

@api_router.post("/hod/summon-student")
async def summon_student(
    summon: SummonStudent,
    token: dict = Depends(verify_token)
):
    """HOD summons a student (creates a notification)"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Verify user is HOD
            cursor.execute("SELECT is_hod, hod_department, name FROM users WHERE id = %s", (token['user_id'],))
            user = cursor.fetchone()
            
            if not user or not user.get('is_hod'):
                raise HTTPException(status_code=403, detail="Only HODs can summon students")
            
            # Get student info
            cursor.execute("SELECT name, department FROM users WHERE id = %s", (summon.student_id,))
            student = cursor.fetchone()
            
            if not student or student['department'] != user['hod_department']:
                raise HTTPException(status_code=404, detail="Student not found in your department")
            
            # Create notification
            cursor.execute("""
                INSERT INTO notifications 
                (user_id, title, message, type, is_read, created_at)
                VALUES (%s, %s, %s, 'summon', FALSE, NOW())
            """, (
                summon.student_id,
                f"Summon from HOD - {user['hod_department']}",
                f"You have been summoned by {user['name']} (HOD). Reason: {summon.reason}" + 
                (f" Time: {summon.scheduled_time}" if summon.scheduled_time else "")
            ))
            conn.commit()
            
            return {"message": f"Summon notification sent to {student['name']}"}

@api_router.get("/notifications")
async def get_notifications(token: dict = Depends(verify_token)):
    """Get notifications for the logged-in user"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM notifications 
                WHERE user_id = %s 
                ORDER BY created_at DESC
                LIMIT 50
            """, (token['user_id'],))
            return cursor.fetchall()

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: int,
    token: dict = Depends(verify_token)
):
    """Mark a notification as read"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                UPDATE notifications SET is_read = TRUE 
                WHERE id = %s AND user_id = %s
            """, (notification_id, token['user_id']))
            conn.commit()
            return {"message": "Notification marked as read"}

# ==========================================
# GEO-FENCED OTP ATTENDANCE
# ==========================================

def calculate_distance(lat1, lon1, lat2, lon2):
    """Calculate Haversine distance between two points in meters"""
    R = 6371000 # Earth radius in meters
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

@api_router.post("/attendance/start-session")
async def start_attendance_session(
    session: AttendanceSessionStart,
    token: dict = Depends(require_role('Teacher'))
):
    """Teacher starts an attendance session"""
    otp = ''.join(random.choices(string.digits, k=6))
    # 60 seconds expiry by default
    expires_at = datetime.now() + timedelta(seconds=60)
    
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Check if course exists and belongs to teacher
            cursor.execute("SELECT * FROM courses WHERE id = %s AND teacher_id = %s", (session.course_id, token['user_id']))
            course = cursor.fetchone()
            if not course:
                raise HTTPException(status_code=403, detail="Unauthorized for this course")
            
            cursor.execute("""
                INSERT INTO attendance_sessions 
                (teacher_id, course_id, otp, lat, lng, radius_meters, expires_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (token['user_id'], session.course_id, otp, session.lat, session.lng, session.radius, expires_at))
            session_id = cursor.lastrowid
            conn.commit()
            
            return {
                "session_id": session_id, 
                "otp": otp, 
                "expires_at": str(expires_at),
                "course_name": course['name']
            }

@api_router.post("/attendance/mark")
async def mark_attendance(
    mark: AttendanceMark,
    token: dict = Depends(require_role('Student'))
):
    """Student marks their attendance with OTP and Geo-fencing"""
    now = datetime.now()
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Find active session by OTP
            cursor.execute("""
                SELECT s.*, c.name as course_name 
                FROM attendance_sessions s
                JOIN courses c ON s.course_id = c.id
                WHERE s.otp = %s AND s.expires_at > %s
            """, (mark.otp, now))
            session = cursor.fetchone()
            
            if not session:
                raise HTTPException(status_code=400, detail="Invalid or expired OTP")
            
            # Verify distance
            dist = calculate_distance(mark.lat, mark.lng, float(session['lat']), float(session['lng']))
            if dist > session['radius_meters']:
                # Record failed attempt notification for teacher
                cursor.execute("""
                    INSERT INTO notifications (user_id, title, message, type)
                    VALUES (%s, 'Radius Violation', 'Student %s attempted to mark attendance for %s from %dm away (Radius: %dm)', 'warning')
                """, (session['teacher_id'], token['user_id'], session['course_name'], int(dist), session['radius_meters']))
                conn.commit()
                raise HTTPException(status_code=400, detail=f"Out of range: {int(dist)}m. Authorized radius is {session['radius_meters']}m.")
            
            # Mark attendance
            try:
                cursor.execute("""
                    INSERT INTO attendance_logs (session_id, student_id, status, marked_at)
                    VALUES (%s, %s, 'present', NOW())
                """, (session['id'], token['user_id']))
                conn.commit()
            except pymysql.err.IntegrityError:
                raise HTTPException(status_code=400, detail="Attendance already marked for this session")
            
            return {"message": "Attendance marked successfully"}

@api_router.get("/attendance/active-sessions")
async def get_active_sessions(token: dict = Depends(verify_token)):
    """Get active sessions (Teacher sees their own, Student sees their relevant ones)"""
    now = datetime.now()
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            if token['role'] == 'Teacher':
                cursor.execute("""
                    SELECT s.*, c.name as course_name, c.code as course_code
                    FROM attendance_sessions s
                    JOIN courses c ON s.course_id = c.id
                    WHERE s.teacher_id = %s AND s.expires_at > %s
                """, (token['user_id'], now))
            else:
                # Get student's department/year
                cursor.execute("SELECT department, year FROM users WHERE id = %s", (token['user_id'],))
                student = cursor.fetchone()
                cursor.execute("""
                    SELECT s.id, s.course_id, c.name as course_name, c.code as course_code, s.expires_at, s.radius_meters
                    FROM attendance_sessions s
                    JOIN courses c ON s.course_id = c.id
                    WHERE c.department = %s AND c.year = %s AND s.expires_at > %s
                """, (student['department'], student['year'], now))
            return cursor.fetchall()

@api_router.get("/attendance/session/{session_id}/logs")
async def get_session_logs(
    session_id: int,
    token: dict = Depends(require_role('Teacher'))
):
    """Teacher views who has marked attendance in a session"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT l.*, u.name as student_name, u.idno as usn
                FROM attendance_logs l
                JOIN users u ON l.student_id = u.id
                WHERE l.session_id = %s
            """, (session_id,))
            return cursor.fetchall()

@api_router.post("/attendance/manual-mark")
async def manual_mark_attendance(
    manual: ManualAttendance,
    token: dict = Depends(require_role('Teacher'))
):
    """Teacher manually marks attendance for a student"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                INSERT INTO attendance_logs (session_id, student_id, status, is_manual, manual_by, marked_at)
                VALUES (%s, %s, %s, TRUE, %s, NOW())
                ON DUPLICATE KEY UPDATE status = VALUES(status), is_manual = TRUE, manual_by = VALUES(manual_by)
            """, (manual.session_id, manual.student_id, manual.status, token['user_id']))
            conn.commit()
            return {"message": "Attendance record updated"}

@api_router.get("/attendance/my-stats")
async def get_my_attendance_stats(token: dict = Depends(require_role('Student'))):
    """Student views their attendance summary by course"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT c.id as course_id, c.name as course_name, c.code as course_code,
                       COUNT(DISTINCT s.id) as total_sessions,
                       COUNT(DISTINCT l.id) as attended_sessions
                FROM courses c
                LEFT JOIN attendance_sessions s ON c.id = s.course_id
                LEFT JOIN attendance_logs l ON s.id = l.session_id AND l.student_id = %s
                WHERE c.department = (SELECT department FROM users WHERE id = %s)
                  AND c.year = (SELECT year FROM users WHERE id = %s)
                GROUP BY c.id
            """, (token['user_id'], token['user_id'], token['user_id']))
            return cursor.fetchall()

@api_router.get("/attendance/all")
async def get_all_attendance(token: dict = Depends(verify_token)):
    """Generic attendance history"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            if token['role'] == 'Teacher':
                cursor.execute("""
                    SELECT l.*, u.name as student_name, u.idno as usn, c.name as course_name, l.marked_at as date
                    FROM attendance_logs l
                    JOIN users u ON l.student_id = u.id
                    JOIN attendance_sessions s ON l.session_id = s.id
                    JOIN courses c ON s.course_id = c.id
                    WHERE s.teacher_id = %s
                    ORDER BY l.marked_at DESC
                """, (token['user_id'],))
            elif token['role'] == 'Student':
                cursor.execute("""
                    SELECT l.*, c.name as course_name, c.code as course_code, l.marked_at as date
                    FROM attendance_logs l
                    JOIN attendance_sessions s ON l.session_id = s.id
                    JOIN courses c ON s.course_id = c.id
                    WHERE l.student_id = %s
                    ORDER BY l.marked_at DESC
                """, (token['user_id'],))
            else: # Admin
                cursor.execute("""
                    SELECT l.*, u.name as student_name, u.idno as usn, c.name as course_name, l.marked_at as date
                    FROM attendance_logs l
                    JOIN users u ON l.student_id = u.id
                    JOIN attendance_sessions s ON l.session_id = s.id
                    JOIN courses c ON s.course_id = c.id
                    ORDER BY l.marked_at DESC
                    LIMIT 500
                """)
            return cursor.fetchall()

# ==========================================
# EXAM HALL LOCATOR
# ==========================================

class ExamCreate(BaseModel):
    name: str
    course_id: int
    exam_date: str
    start_time: str
    end_time: str
    is_visible: bool = False

class ExamHallCreate(BaseModel):
    name: str
    building: str
    floor: int
    capacity: int

class GenerateSeating(BaseModel):
    exam_id: int
    hall_ids: List[int]

@api_router.post("/exams")
async def create_exam(
    exam: ExamCreate,
    token: dict = Depends(require_role('Admin'))
):
    """Admin creates an exam"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                INSERT INTO exam_schedules 
                (name, course_id, exam_date, start_time, end_time, is_visible, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, NOW())
            """, (exam.name, exam.course_id, exam.exam_date, exam.start_time, 
                  exam.end_time, exam.is_visible))
            conn.commit()
            return {"message": "Exam created", "id": cursor.lastrowid}

@api_router.get("/exams")
async def get_exams(
    token: dict = Depends(require_role('Admin'))
):
    """Admin gets all exams"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT e.*, c.name as course_name, c.code as course_code
                FROM exam_schedules e
                LEFT JOIN courses c ON e.course_id = c.id
                ORDER BY e.exam_date DESC
            """)
            return cursor.fetchall()

@api_router.put("/exams/{exam_id}/toggle-visibility")
async def toggle_exam_visibility(
    exam_id: int,
    token: dict = Depends(require_role('Admin'))
):
    """Toggle exam visibility for students"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                UPDATE exam_schedules SET is_visible = NOT is_visible WHERE id = %s
            """, (exam_id,))
            conn.commit()
            return {"message": "Visibility toggled"}

@api_router.post("/exams/halls")
async def create_exam_hall(
    hall: ExamHallCreate,
    token: dict = Depends(require_role('Admin'))
):
    """Admin creates an exam hall"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                INSERT INTO exam_halls (name, building, floor, capacity)
                VALUES (%s, %s, %s, %s)
            """, (hall.name, hall.building, hall.floor, hall.capacity))
            conn.commit()
            return {"message": "Hall created", "id": cursor.lastrowid}

@api_router.get("/exams/halls")
async def get_exam_halls(token: dict = Depends(require_role('Admin'))):
    """Get all exam halls"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM exam_halls ORDER BY building, floor, name")
            return cursor.fetchall()

@api_router.post("/exams/{exam_id}/generate-seating")
async def generate_seating_arrangement(
    exam_id: int,
    seating: GenerateSeating,
    token: dict = Depends(require_role('Admin'))
):
    """Generate seating arrangement with department interleaving"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            # Get exam details
            cursor.execute("SELECT course_id FROM exam_schedules WHERE id = %s", (exam_id,))
            exam = cursor.fetchone()
            if not exam:
                raise HTTPException(status_code=404, detail="Exam not found")
            
            # Get course department
            cursor.execute("SELECT department FROM courses WHERE id = %s", (exam['course_id'],))
            course = cursor.fetchone()
            
            # Get all students for this exam (can be multiple departments)
            cursor.execute("""
                SELECT u.id, u.name, u.idno as usn, u.department
                FROM users u
                WHERE u.role = 'Student'
                ORDER BY u.department, u.name
            """)
            students = cursor.fetchall()
            
            # Get halls with capacities
            cursor.execute("""
                SELECT * FROM exam_halls WHERE id IN %s ORDER BY building, floor
            """, (tuple(seating.hall_ids),))
            halls = cursor.fetchall()
            
            if not halls:
                raise HTTPException(status_code=400, detail="No halls selected")
            
            total_capacity = sum(h['capacity'] for h in halls)
            if len(students) > total_capacity:
                raise HTTPException(status_code=400, 
                    detail=f"Not enough seats. Students: {len(students)}, Capacity: {total_capacity}")
            
            # Interleave students by department
            departments = {}
            for s in students:
                dept = s['department'] or 'Unknown'
                if dept not in departments:
                    departments[dept] = []
                departments[dept].append(s)
            
            # Interleave
            interleaved = []
            dept_lists = list(departments.values())
            max_len = max(len(d) for d in dept_lists) if dept_lists else 0
            for i in range(max_len):
                for dept_list in dept_lists:
                    if i < len(dept_list):
                        interleaved.append(dept_list[i])
            
            # Clear existing seating for this exam
            cursor.execute("DELETE FROM exam_seating WHERE exam_id = %s", (exam_id,))
            
            # Assign seats
            seat_number = 1
            hall_idx = 0
            hall_seat_count = 0
            
            for student in interleaved:
                if hall_idx >= len(halls):
                    break
                    
                current_hall = halls[hall_idx]
                
                cursor.execute("""
                    INSERT INTO exam_seating 
                    (exam_id, student_id, hall_id, seat_number, row_number)
                    VALUES (%s, %s, %s, %s, %s)
                """, (exam_id, student['id'], current_hall['id'], 
                      seat_number, hall_seat_count // 10 + 1))
                
                hall_seat_count += 1
                seat_number += 1
                
                if hall_seat_count >= current_hall['capacity']:
                    hall_idx += 1
                    hall_seat_count = 0
            
            conn.commit()
            return {"message": f"Seating generated for {len(interleaved)} students"}

@api_router.get("/exams/{exam_id}/seating")
async def get_exam_seating(
    exam_id: int,
    token: dict = Depends(require_role('Admin'))
):
    """Get seating arrangement for an exam"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT es.*, u.name as student_name, u.idno as usn, u.department,
                       eh.name as hall_name, eh.building, eh.floor
                FROM exam_seating es
                JOIN users u ON es.student_id = u.id
                JOIN exam_halls eh ON es.hall_id = eh.id
                WHERE es.exam_id = %s
                ORDER BY eh.building, eh.floor, es.seat_number
            """, (exam_id,))
            return cursor.fetchall()

@api_router.get("/exams/my-seat")
async def get_my_exam_seat(token: dict = Depends(verify_token)):
    """Student gets their exam seat"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT es.*, e.name as exam_name, e.exam_date, e.start_time, e.end_time,
                       c.name as course_name, c.code as course_code,
                       eh.name as hall_name, eh.building, eh.floor
                FROM exam_seating es
                JOIN exam_schedules e ON es.exam_id = e.id
                JOIN courses c ON e.course_id = c.id
                JOIN exam_halls eh ON es.hall_id = eh.id
                WHERE es.student_id = %s AND e.is_visible = TRUE
                  AND e.exam_date >= CURDATE()
                ORDER BY e.exam_date, e.start_time
            """, (token['user_id'],))
            return cursor.fetchall()

@api_router.get("/exams/upcoming")
async def get_upcoming_exams(token: dict = Depends(verify_token)):
    """Get upcoming visible exams for students"""
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT e.*, c.name as course_name, c.code as course_code
                FROM exam_schedules e
                JOIN courses c ON e.course_id = c.id
                WHERE e.is_visible = TRUE AND e.exam_date >= CURDATE()
                ORDER BY e.exam_date, e.start_time
            """)
            return cursor.fetchall()

# Include the router
app.include_router(api_router)



logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
